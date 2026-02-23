import streamlit as st
import polars as pl
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

st.set_page_config(page_title="Xử lý data kiểm date", layout="wide")

st.title("📊 Xử lý dữ liệu kiểm date")
st.write("Upload folder chứa các file Excel → xuất file tổng hợp")

# =============================

# CONFIG

# =============================

COLUMNS_KEEP = [
'Mã siêu thị','Tên siêu thị','Mã sản phẩm','Tên sản phẩm',
'SL chuyển kho','SL giảm giá','SL hủy tại siêu thị',
'Số lượng trả NCC','SL đổi hàng NCC','Số lượng bình thường',
'SL tặng KM','SL cận date (tặng quà)',
'Ngày tạo','Lần kiểm cuối cùng',
'Mã nhân viên','Họ và tên nhân viên',
'Ngày duyệt','Người duyệt','Tên người duyệt',
'Hình ảnh','Ghi chú trạng thái','Ghi chú',
'Ngày hệ thống yêu cầu','Trạng thái','Nội dung',
'Hạn sử dụng','Date gần nhất','Hình ảnh_1',
'Phân loại','Thời gian bắt đầu','Thời gian kết thúc',
'Giá trị phần trăm giảm giá'
]

# =============================

# DATA PROCESS FUNCTION

# =============================

def process_dataframe(df: pl.DataFrame) -> pl.DataFrame:

    
    df = df.with_columns([
        pl.col('Hình ảnh_1').str.replace_all('^(.*)$', '"$1"')
    ])
    
    df = df.with_columns([
        pl.col('SL giảm giá').cast(pl.Float64),
        pl.col('SL hủy tại siêu thị').cast(pl.Float64),
        pl.col('SL tặng KM').cast(pl.Float64),
        pl.col('SL cận date (tặng quà)').cast(pl.Float64)
    ])
    
    df = df.with_columns([
        (
            pl.col('SL giảm giá') +
            pl.col('SL hủy tại siêu thị') +
            pl.col('SL tặng KM') +
            pl.col('SL cận date (tặng quà)')
        ).alias('Điều kiện lọc')
    ])
    
    df = df.filter(pl.col('Điều kiện lọc') > 0)
    
    df = df.select(COLUMNS_KEEP)
    
    return df


# =============================

# READ & PROCESS ONE FILE

# =============================

def concat_excel_sheets(file_bytes, file_name):

    
    def get_sheet_names(excel_io):
        wb = load_workbook(excel_io, read_only=True)
        return wb.sheetnames
    
    def read_sheet(sheet_name):
    
        df = pl.read_excel(file_bytes, sheet_name=sheet_name)
    
        # ép kiểu string để tránh lỗi schema
        df = df.with_columns(
            [pl.col(c).cast(pl.Utf8) for c in df.columns]
        )
    
        # xử lý từng sheet
        df = process_dataframe(df)
    
        # thêm tên file nguồn
        df = df.with_columns(
            pl.lit(file_name).alias("file_name")
        )
    
        return df
    
    excel_io = BytesIO(file_bytes.getvalue())
    sheet_names = get_sheet_names(excel_io)
    
    max_workers = min(4, len(sheet_names))
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        dfs = list(executor.map(read_sheet, sheet_names))
    
    result = pl.concat(dfs, how="diagonal_relaxed")
    
    return result


# =============================

# READ MULTIPLE FILES

# =============================

def read_excel_files(uploaded_files):

    
    all_dfs = []
    
    for file in uploaded_files:
        try:
            df = concat_excel_sheets(file, file.name)
            all_dfs.append(df)
        except Exception as e:
            st.error(f"Lỗi file {file.name}: {e}")
    
    if not all_dfs:
        return None
    
    final = pl.concat(all_dfs, how="diagonal_relaxed")
    
    return final


# =============================

# UI

# =============================

uploaded_files = st.file_uploader(
"📂 Upload các file Excel trong folder",
type=["xlsx", "xls"],
accept_multiple_files=True
)

if uploaded_files:

    
    st.success(f"Đã upload {len(uploaded_files)} file")
    
    if st.button("🚀 Xử lý dữ liệu"):
    
        with st.spinner("Đang xử lý..."):
    
            data_date = read_excel_files(uploaded_files)
    
            if data_date is None:
                st.error("Không đọc được dữ liệu")
                st.stop()
    
            output = BytesIO()
    
            data_date.to_pandas().to_excel(
                output,
                index=False,
                engine="xlsxwriter"
            )
    
            output.seek(0)
    
            st.success("✅ Hoàn thành!")
    
            st.download_button(
                label="📥 Download file Excel",
                data=output,
                file_name="data_kiem_date.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

